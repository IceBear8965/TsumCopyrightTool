"""
Copyright (C) 2025 IceBear8965

This program is free software: you can redistribute it and/or
modify it under the terms of the GNU General Public License as published
by the Free Software Foundation, either version 3 of the License, or (at your option)
any later version.

This program is distributed in the hope that it will be useful, but WITHOUT ANY
WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR
PURPOSE. See the GNU General Public License for more details.
"""
import os
from pathlib import Path

import openpyxl
from PyQt5.QtCore import Q_ARG, QMetaObject, QObject, Qt, QThread, pyqtSignal, pyqtSlot, QSize
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QWidget
from PyQt5.QtGui import QIcon
from qfluentwidgets import InfoBar, InfoBarPosition, RadioButton, getIconColor, Theme

from app.common.addDots import addDots
from app.common.presetModel import presetModel
from app.common.config import cfg
from app.common.excelHandler import excelHandler
from app.common.icon import CustomIcons
from app.common.parsers.arenaParser import parseArena
from app.common.parsers.kidisParser import parseKidis
from app.common.parsers.saksParser import parseSaks
from app.common.parsers.sauconyParser import parseSaucony
from app.common.saver import Saver
from app.common.setting import DOWNLOAD_FOLDER
from app.common.sortInput import sortInput
from app.components.file_card import FileCard
from app.view.excel_interface.UI_ExcelInterface import Ui_ExcelInterface
from app.common.set_websites_names import set_websites_names


# Обработчик таблиц в 2 потоке
class ExcelParser(QObject):
    format_result = pyqtSignal(list)
    parse_result = pyqtSignal(list)
    progress_signal = pyqtSignal(int)

    @pyqtSlot(list, int, list, list, str)
    def parseExcelInThread(self, values, columnIndex, filters, order, websiteName):
        sources = []
        for value in values:
            try:
                sources.append(value[columnIndex])
            except IndexError:
                sources.append(None)

        data = []
        i = 1
        for source in sources:
            match websiteName:
                case "Saks85":
                    try:
                        parsedData = parseSaks(source, filters, order)
                    except Exception:
                        parsedData = source  # Ретёрним линк, если страница не существует
                case "Saucony":
                    try:
                        parsedData = parseSaucony(source)
                    except Exception:
                        parsedData = source  # Ретёрним линк, если страница не существует
                case "Arena":
                    try:
                        parsedData = parseArena(source)
                    except Exception:
                        parsedData = source  # Ретёрним линк, если страница не существует
                case "Kidis":
                    try:
                        parsedData = parseKidis(source, filters, order)
                    except Exception:
                        parsedData = source  # Ретёрним линк, если страница не существует
            data.append(parsedData)
            self.progress_signal.emit(i)
            i += 1

        self.parse_result.emit(data)

    @pyqtSlot(list, int, list, list)
    def formatExcelInThread(self, values, columnIndex, filters, order):
        data = []
        for value in values:
            try:
                data.append(value[columnIndex])
            except IndexError:
                data.append(None)

        output = []
        i = 1
        for value in data:
            try:
                sorted = sortInput(value.split("\n"), filters, order)
                dotted = addDots(sorted)
                output.append(dotted)
            except Exception:
                output.append(value)
            self.progress_signal.emit(i)
            i += 1

        self.format_result.emit(output)


class ExcelInterface(Ui_ExcelInterface, QWidget):
    sheetChanged = pyqtSignal(list, int, int)

    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setupUi(self)
        self.current_filename = ""
        self.current_workbook = None
        self.current_sheetname = ""
        self.current_values = []
        self.max_row = 0

        self.setObjectName("excelInterface")
        self.toggleUrlParsing()
        self.tablePreview.verticalHeader().show()
        self.useUrlToggle.setOnText("Parsing from url`s")
        self.useUrlToggle.setOffText("Formatting descriptions")
        self.progressBar.setProperty("value", 0)
        self.sheetsView.enableTransparentBackground()
        self.saver = Saver()

        # Multithreading
        self.ExcelParser = ExcelParser()
        self.ExcelThread = QThread()
        self.ExcelParser.parse_result.connect(self.on_parse_excel_result_ready)
        self.ExcelParser.format_result.connect(self.on_format_excel_result_ready)
        self.ExcelParser.progress_signal.connect(self.progressUpdate)
        self.ExcelParser.moveToThread(self.ExcelThread)
        self.ExcelThread.start()

        self.fileCard = FileCard(CustomIcons.XLSX, "Input Excel file", "", self)  # Карточка выбора эксель файла
        self.inputFileCard.addWidget(self.fileCard)

        # Иконки при выборе имени сайта
        set_websites_names(self.websiteNameCombo)  # Добавляет элементы выбора в комбо бокс

        # connect signal to slots
        self.useUrlToggle.checkedChanged.connect(self.toggleUrlParsing)
        self.fileCard.openButton.clicked.connect(self.getExcelFile)
        self.excelRunBtn.clicked.connect(self.processExcel)
        self.sheetChanged.connect(self.loadExcelTable)

    # Интерфейс
    def toggleUrlParsing(self):  # Сбрасываем выбор парсинга и блокируем выбор сайта
        if self.useUrlToggle.isChecked():
            self.websiteNameCombo.setEnabled(True)
            self.excelRunBtn.setText("Parse")
        else:
            self.websiteNameCombo.setEnabled(False)
            self.excelRunBtn.setText("Format")

    def getExcelFile(self):
        excel_file, _ = QFileDialog.getOpenFileName(self, "Open file", str(DOWNLOAD_FOLDER), "Excel files (*.xlsx)")
        if excel_file != "":
            self.onNewFileLoaded()
            self.fileCard.setContent(excel_file)
            self.tablePreview.clearSelection()
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb.worksheets[0]
            values, max_row, max_column = excelHandler.getData(ws)
            self.loadExcelTable(values, max_row, max_column)
            self.setSheets(wb)
            self.current_filename = Path(excel_file).stem
            self.current_workbook = wb
            self.current_sheetname = wb.sheetnames[0]
            self.current_values = values
            self.max_row = max_row
        else:
            InfoBar.warning(
                title="Error",
                content="Choose Excel file",
                orient=Qt.Horizontal,
                isClosable=True,
                duration=2000,
                position=InfoBarPosition.BOTTOM_RIGHT,
                parent=self,
            )

    def setSheets(self, workbook):
        sheetNames = workbook.sheetnames
        for sheet in sheetNames:
            btn = RadioButton(sheet)
            btn.sheet = sheet
            btn.clicked.connect(self.sheetToggled)
            self.horizontalLayout_3.addWidget(btn, 0, Qt.AlignTop)
        first_btn = self.horizontalLayout_3.itemAt(0).widget()
        first_btn.setChecked(True)

        # horizontalLayout_3 -- лейаут в scrollArea, хрен знает как переименовать его из дизайнера

    def loadExcelTable(self, values, max_row, max_column):
        self.tablePreview.clear()
        self.current_values = values
        self.max_row = max_row
        self.tablePreview.setRowCount(max_row)
        self.tablePreview.setColumnCount(max_column)
        for i in range(max_column):
            self.tablePreview.setColumnWidth(i, 170)
        self.tablePreview.horizontalHeader().hide()

        rowIndex = 0
        for row in values:
            colIndex = 0
            for column in row:
                try:
                    self.tablePreview.setItem(rowIndex, colIndex, QTableWidgetItem(column))
                except Exception:
                    self.tablePreview.setItem(rowIndex, colIndex, QTableWidgetItem(None))
                colIndex += 1
            rowIndex += 1

    def sheetToggled(self):
        radiobutton = self.sender()
        if radiobutton.isChecked():
            current_sheet = self.current_workbook[radiobutton.sheet]
            self.current_sheetname = radiobutton.sheet
            values, max_row, max_column = excelHandler.getData(current_sheet)
            self.sheetChanged.emit(values, max_row, max_column)

    def toggleSheetSelection(self):  # Отключаем смену страницы при обработке страниц
        for btn in reversed(range(self.horizontalLayout_3.count())):
            btn_widget_state = self.horizontalLayout_3.itemAt(btn).widget().isEnabled()
            self.horizontalLayout_3.itemAt(btn).widget().setEnabled(not btn_widget_state)

    def onNewFileLoaded(self):  # удаляем все листы из панели выбора
        for i in reversed(range(self.horizontalLayout_3.count())):
            self.horizontalLayout_3.itemAt(i).widget().setParent(None)

    # Write data to Excel file
    def processExcel(self):
        column = self.tablePreview.selectedItems()
        websiteName = self.websiteNameCombo.currentText()
        columnIndex = None
        if len(column) > 0:
            columnIndex = column[0].column()
        filters, paramorder = presetModel.getSetting()

        # Отправка в поток-обработчик
        if isinstance(columnIndex, int) and self.current_values != []:
            if self.useUrlToggle.isChecked():  # парсинг
                self.fileCard.openButton.setEnabled(False)
                self.excelRunBtn.setEnabled(False)
                self.useUrlToggle.setEnabled(False)
                self.websiteNameCombo.setEnabled(False)
                self.toggleSheetSelection()
                QMetaObject.invokeMethod(  # вкидываем данные в поток
                    self.ExcelParser,
                    "parseExcelInThread",
                    Qt.ConnectionType.QueuedConnection,
                    Q_ARG(list, self.current_values),
                    Q_ARG(int, columnIndex),
                    Q_ARG(list, filters),
                    Q_ARG(list, paramorder),
                    Q_ARG(str, websiteName),
                )
                self.progressBar.setValue(0)
                self.progressBar.setRange(0, self.max_row)
                InfoBar.success(
                    title="Parsing",
                    content="Parsing your table",
                    orient=Qt.Horizontal,
                    isClosable=True,
                    duration=2000,
                    position=InfoBarPosition.TOP_RIGHT,
                    parent=self,
                )
            else:  # форматирование
                self.fileCard.openButton.setEnabled(False)
                self.excelRunBtn.setEnabled(False)
                self.useUrlToggle.setEnabled(False)
                self.toggleSheetSelection()
                QMetaObject.invokeMethod(  # вкидываем данные в поток
                    self.ExcelParser,
                    "formatExcelInThread",
                    Qt.ConnectionType.QueuedConnection,
                    Q_ARG(list, self.current_values),
                    Q_ARG(int, columnIndex),
                    Q_ARG(list, filters),
                    Q_ARG(list, paramorder)
                )
                self.progressBar.setValue(0)
                self.progressBar.setRange(0, self.max_row)
                InfoBar.success(
                    title="Formatting",
                    content="Formatting your table",
                    orient=Qt.Horizontal,
                    isClosable=True,
                    duration=2000,
                    position=InfoBarPosition.TOP_RIGHT,
                    parent=self,
                )
        else:
            InfoBar.warning(
                title="Error",
                content="Select file and select column in preview",
                orient=Qt.Horizontal,
                isClosable=True,
                duration=2000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )

    # Multithreading get results from ExcelParser worker
    def on_parse_excel_result_ready(self, output):
        filename = f"{self.current_filename}-{self.current_sheetname}-output.xlsx"
        output_file = os.path.join(cfg.get(cfg.outputFolder), filename)
        try:
            self.saver.saveToExcel(output, output_file)
            InfoBar.success(
                title="Parsing",
                content="Excel table parsed successfully",
                orient=Qt.Horizontal,
                isClosable=True,
                duration=2000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )

        except Exception:
            InfoBar.error(
                title="Saving",
                content="An error occurred while writing the file. Check that it is not open in another program.",
                orient=Qt.Horizontal,
                isClosable=True,
                duration=2000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )

        self.fileCard.openButton.setEnabled(True)
        self.excelRunBtn.setEnabled(True)
        self.useUrlToggle.setEnabled(True)
        self.websiteNameCombo.setEnabled(True)
        self.toggleSheetSelection()

    def on_format_excel_result_ready(self, output):
        filename = f"{self.current_filename}-{self.current_sheetname}-output.xlsx"
        output_file = os.path.join(cfg.get(cfg.outputFolder), filename)
        try:
            self.saver.saveToExcel(output, output_file)

            InfoBar.success(
                title="Formatting",
                content="Excel table formatted successfully",
                orient=Qt.Horizontal,
                isClosable=True,
                duration=2000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )
        except Exception:
            InfoBar.error(
                title="Saving",
                content="An error occurred while writing the file. Check that it is not open in another program.",
                orient=Qt.Horizontal,
                isClosable=True,
                duration=2000,
                position=InfoBarPosition.TOP_RIGHT,
                parent=self,
            )

        self.fileCard.openButton.setEnabled(True)
        self.excelRunBtn.setEnabled(True)
        self.useUrlToggle.setEnabled(True)
        self.toggleSheetSelection()

    def progressUpdate(self, value):
        self.progressBar.setValue(value)